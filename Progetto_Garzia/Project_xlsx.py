from datetime import datetime
print(datetime.today().strftime('%Y-%m-%d'))

import openpyxl
data_transfer = openpyxl.load_workbook(r'.\\Corsi\Dati\Documenti\Corsisti\PPBD01-10\My Documents\Corso_Python\PPBD01\Progetto_Garzia')
foglio_candidati= data_transfer ['Candidati']
lista_nomi_fogli= data_transfer.sheetnames
print(lista_nomi_fogli)


for row in range(1, rows + 1):
    string = ''
    for column in range(1, cols + 1):
        value = sheet.cell(row+1, column = column).value
        string = string + str(value) + ', '
    string = string[:-2]
    check = sheet.cell(row+1, column=20)
    check = str(check.value)
    if check == 'CV al Cliente':    
        print(foglio_candidati)
