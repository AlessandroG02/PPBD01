import imaplib
import email
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
import os

class DataReciver:
    @staticmethod
    def reciver_data():
        username = 'momo_edo_python@outlook.com'
        password = '!Torino23!'

        # Connettiti al server IMAP di Outlook
        mail = imaplib.IMAP4_SSL('Outlook.office365.com', 993)

        # Effettua il login
        mail.login(username, password)

        # Seleziona la casella di posta in arrivo
        mail.select('inbox')

        # Cerca le email nella casella di posta in arrivo
        result, data = mail.search(None, 'ALL')

        # Otteniamo l'ID dell'ultima email ricevuta
        email_ids = data[0].split()
        latest_email_id = email_ids[-1]

        # Otteniamo il corpo dell'email
        result, email_data = mail.fetch(latest_email_id, '(RFC822)')
        raw_email = email_data[0][1]

        # Parsing del contenuto dell'email
        msg = email.message_from_bytes(raw_email)

        # Estrai il contenuto del corpo del messaggio
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == "text/html":
                    body = part.get_payload(decode=True).decode('utf-8')
                    break
        else:
            body = msg.get_payload(decode=True).decode('utf-8')

        # Chiudi la connessione con il server IMAP
        mail.logout()

        # Effettua il parsing dell'HTML
        soup = BeautifulSoup(body, 'html.parser')

        # Trova la tabella con il tag specificato
        table = soup.find('table', {'border': '0', 'cellspacing': '0', 'cellpadding': '0', 'width': '1313'})

        # Estrai i dati dalla tabella
        data = []
        for row in table.find_all('tr'):
            row_data = []
            for cell in row.find_all('td'):
                row_data.append(cell.get_text(strip=True))
            data.append(row_data)

        # Crea il DataFrame dal file di input
        df_new = pd.DataFrame(data)

        # Rinomina le colonne
        df_new.columns = df_new.iloc[0]
        df_new = df_new[1:]  # Rimuovi la riga dei nomi delle colonne

        # Converti la colonna 'Id ricerca' in tipo numerico
        df_new['Id ricerca'] = pd.to_numeric(df_new['Id ricerca'], errors='coerce')

        # Ordina il DataFrame in base alla colonna 'Id ricerca'
        df_new = df_new.sort_values(by='Id ricerca')

        # Converti nuovamente la colonna 'Id ricerca' in tipo stringa
        df_new['Id ricerca'] = df_new['Id ricerca'].astype(str)

        # Percorso del file di input e output
        input_file_path = 'C:/Users/Utente/Documents/Progetto_Garzia/ACE10001I C-Lab HR/DB C-Lab (HRR).xlsx'
        output_file_path = 'C:/Users/Utente/Documents/Progetto_Garzia/ACE10001I C-Lab HR/DB C-Lab (HRR)_output.xlsx'

        # Verifica se il file di output esiste già
        if os.path.exists(output_file_path):
            # Carica il workbook di output
            workbook = load_workbook(output_file_path)

            # Seleziona il worksheet
            worksheet = workbook['Richieste']

            # Leggi il DataFrame dal file di output
            output_df = pd.DataFrame(worksheet.values)

            # Rinomina le colonne del DataFrame di output
            output_df.columns = output_df.iloc[0]
            output_df = output_df[1:]  # Rimuovi la riga dei nomi delle colonne

            # Unisci i dati esistenti con i nuovi dati
            merged_data = pd.concat([output_df, df_new], ignore_index=True)

            # Rimuovi gli "Id ricerca" duplicati
            merged_data = merged_data.drop_duplicates(subset='Id ricerca', keep='first')

            # Rimuovi il worksheet esistente
            workbook.remove(worksheet)
        else:
            # Carica un nuovo workbook
            workbook = Workbook()
            merged_data = df_new

        # Crea un nuovo worksheet con i dati combinati
        worksheet = workbook.create_sheet('Richieste')

        # Scrivi i dati del DataFrame di input nel worksheet
        input_df = pd.read_excel(input_file_path, sheet_name='Richieste')
        for row in dataframe_to_rows(input_df, index=False, header=True):
            worksheet.append(row)

        # Scrivi i dati del DataFrame combinato nel worksheet
        for row in dataframe_to_rows(merged_data, index=False, header=False):
            worksheet.append(row)

        # Salva il workbook modificato
        workbook.save(output_file_path)

        print("Dati aggiunti correttamente.")
