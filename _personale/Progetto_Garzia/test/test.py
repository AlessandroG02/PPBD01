import imaplib
import email
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Configurazione delle credenziali IMAP
username = 'momo_edo_python@outlook.com'
password = '!Torino23!'

# Connettiti al server IMAP di Outlook
mail = imaplib.IMAP4_SSL('Outlook.office365.com', 993)

# Effettua il login
mail.login(username, password)

# Seleziona la casella di posta in arrivo
mail.select('inbox')

# Cerca le email nella casella di posta in arrivo
result, data = mail.search(None, 'ALL')

# Otteniamo l'ID dell'ultima email ricevuta
email_ids = data[0].split()
latest_email_id = email_ids[-1]

# Otteniamo il corpo dell'email
result, email_data = mail.fetch(latest_email_id, '(RFC822)')
raw_email = email_data[0][1]

# Parsing del contenuto dell'email
msg = email.message_from_bytes(raw_email)

# Estrai il contenuto del corpo del messaggio
if msg.is_multipart():
    for part in msg.walk():
        if part.get_content_type() == "text/html":
            body = part.get_payload(decode=True).decode('utf-8')
            break
else:
    body = msg.get_payload(decode=True).decode('utf-8')

# Chiudi la connessione con il server IMAP
mail.logout()

# Effettua il parsing dell'HTML
soup = BeautifulSoup(body, 'html.parser')

# Trova la tabella con il tag specificato
table = soup.find('table', {'border': '0', 'cellspacing': '0', 'cellpadding': '0', 'width': '1313'})

# Estrai i dati dalla tabella
data = []
for row in table.find_all('tr'):
    row_data = []
    for cell in row.find_all('td'):
        row_data.append(cell.get_text(strip=True))
    data.append(row_data)

# Crea il DataFrame dal file di input
df = pd.DataFrame(data)

# Rinomina le colonne
df.columns = df.iloc[0]
df = df[1:]  # Rimuovi la riga dei nomi delle colonne

# Ordina il DataFrame in base alla colonna 'Id ricerca'
df = df.sort_values(by='Id ricerca')

# Rimuovi gli "Id ricerca" duplicati
df = df.drop_duplicates(subset='Id ricerca', keep='first')

# Percorso del file di input
input_file_path = '/ACE10001I C-Lab HR/DB C-Lab (HRR).xlsx'

# Carica il workbook di input
workbook = load_workbook(input_file_path)

# Seleziona il worksheet "Richieste"
worksheet = workbook['Richieste']

# Ottieni l'ultima riga del worksheet
last_row = worksheet.max_row + 1

# Aggiungi le righe del DataFrame al file di input
for row in dataframe_to_rows(df, index=False, header=True):
    worksheet.append(row)

# Salva il workbook modificato
workbook.save(input_file_path)

print("Dati aggiunti correttamente al file DB C-Lab (HRR).xlsx nella sheet Richieste.")
