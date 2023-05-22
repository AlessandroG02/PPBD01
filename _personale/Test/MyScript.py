import datetime
import os
import re
import shutil
import openpyxl
import zipfile
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# 1 Step Creazione del file di output
book_transfer = openpyxl.Workbook()
foglio_transfer = book_transfer.active
foglio_transfer.title = 'Candidati'

# Salvataggio del file di output
book_transfer.save('./DB C-Lab (Transfer).xlsx')

# Apertura del file di input
book = openpyxl.load_workbook('DB C-Lab (HRR).xlsx', data_only=True)
target_sheet = 'Candidati'

# Controllo che il foglio "Candidati" esista nel file di input
if target_sheet not in book.sheetnames:
    raise ValueError(f'Il foglio "{target_sheet}" non è stato trovato nella cartella di lavoro.')

foglio_candidati = book[target_sheet]

# Individuazione del numero di colonna della data oggetto del filtro
target_col_label = 'Data invio CV al cliente'
target_col = None

for col_num in range(1, foglio_candidati.max_column + 1):
    column_name = foglio_candidati.cell(1, col_num).value
    if column_name == target_col_label:
        target_col = col_num
        break

# Controllo che la colonna con la data oggetto del filtro esista
if not target_col:
    raise ValueError(f'La colonna "{target_col_label}" non è stata trovata nel foglio "{target_sheet}".')

# Filtraggio dei candidati
target_date = datetime.date.fromisoformat('2023-05-15')
candidati_filtrati = []

for row_num in range(1, foglio_candidati.max_row + 1):
    date_cell = foglio_candidati.cell(row_num, target_col).value
    if row_num == 1 or (isinstance(date_cell, datetime.datetime) and date_cell.date() == target_date):
        record = []
        for col_num in range(1, foglio_candidati.max_column + 1):
            record.append(foglio_candidati.cell(row_num, col_num).value)
        candidati_filtrati.append(record)

# Scrittura dei dati nel foglio di output 'Candidati'
for row_num, row in enumerate(candidati_filtrati):
    for col_num, cell in enumerate(row):
        if isinstance(cell, datetime.datetime):
            cell = cell.strftime('%d/%m/%Y')
        foglio_transfer.cell(row=row_num + 1, column=col_num + 1).value = cell

# Salvataggio del file di output
book_transfer.save('./DB C-Lab (Transfer).xlsx')

# 2 Secondo Step Skill
# Apertura del file di output
book_output = openpyxl.load_workbook('DB C-Lab (Transfer).xlsx', data_only=True)
foglio_output = book_output['Candidati']

# Apertura del file di input
book_input = openpyxl.load_workbook('DB C-Lab (HRR).xlsx', data_only=True)
foglio_input = book_input['AnagSkill']

# Estrazione degli ID candidati numerici dal foglio di output
colonna_id_candidato = 'H'  # Assumiamo che la colonna "Id candidato" sia la colonna H, da modificare se necessario
id_candidati = set()

for row in foglio_output.iter_rows(values_only=True):
    id_candidato = row[foglio_output[colonna_id_candidato + '1'].column - 1]
    if isinstance(id_candidato, int):
        id_candidati.add(id_candidato)

# Copia della prima riga nel foglio di output
prima_riga_input = foglio_input[1]
foglio_output_copia = book_output.create_sheet(title='AnagSkill', index=0)
prima_riga_output = foglio_output_copia.append([cell.value for cell in prima_riga_input])

# Confronto e copia delle righe nel foglio di output
for row_input in foglio_input.iter_rows(min_row=2, values_only=True):
    progr_interno = row_input[1]
    # Assumiamo che la colonna "Progr Interno" sia la colonna B, da modificare se necessario

    if isinstance(progr_interno, int) and progr_interno in id_candidati:
        foglio_output_copia.append(row_input)

# Salvataggio del file di output
book_output.save('DB C-Lab (Transfer).xlsx')

# 3 Terzo Step PDF

# Definizione delle directory di input e output
directory_cv_cliente = './CV al Cliente'
directory_invio_email = './Invio Email'

# Apertura del file di output
book_output = openpyxl.load_workbook('DB C-Lab (Transfer).xlsx', data_only=True)
foglio_output = book_output['AnagSkill']

# Ottenere la colonna 'E' (Cognome) e la colonna 'D' (Nome) dal foglio di output
colonna_cognome = foglio_output['E']
colonna_nome = foglio_output['D']

# Iterazione sulle righe del foglio di output
for row_output in foglio_output.iter_rows(min_row=2, values_only=True):
    cognome = row_output[3]  # Assumiamo che la colonna "E" sia la colonna 4, da modificare se necessario
    nome = row_output[4]  # Assumiamo che la colonna "D" sia la colonna 3, da modificare se necessario

    # Rimozione degli spazi in eccesso nel nome e cognome
    cognome = cognome.replace(" ", "") if cognome else ""
    nome = nome.strip() if nome else ""

    print("Cognome:", cognome)
    print("Nome:", nome)

    if cognome and nome:  # Verifica se il valore delle celle non è vuoto
        nome_cognome = f'{cognome}{nome}'.lower()
        files_cv = os.listdir(directory_cv_cliente)

        for file_cv in files_cv:
            file_cv_normalized = re.sub(r'\W+', '',
                                        file_cv).lower()  # Rimuove caratteri non alfanumerici e converte in minuscolo
            if nome_cognome in file_cv_normalized:
                percorso_file_pdf = os.path.join(directory_cv_cliente, file_cv)
                percorso_destinazione = os.path.join(directory_invio_email, file_cv)
                shutil.copy2(percorso_file_pdf, percorso_destinazione)
                print("File copiato:", percorso_destinazione)
                break
        else:
            print("File PDF non trovato per il nome e cognome:", cognome, nome)

# Salvataggio del file di output
book_output.save('DB C-Lab (Transfer).xlsx')

# Copia del file di output nella directory 'Invio Email'
percorso_file_output = 'DB C-Lab (Transfer).xlsx'
percorso_destinazione_output = './Invio Email/DB C-Lab (Transfer).xlsx'
shutil.copy2(percorso_file_output, percorso_destinazione_output)
print("File di output copiato:", percorso_destinazione_output)

# 4 e 5 Step  Zip e invio Email

# Definisci i dettagli dell'email
mittente = "mohmedelharf90@gmail.com"
password = input("Inserisci la password: ")
destinatario = "torfis.developer@gmail.com"
oggetto = "Invio File"
testo = "Ciao, allego il file compresso della directory 'Invio Email'."

# Comprimi la directory 'Invio Email' in un file zip
nome_zip = "Invio_Email.zip"
with zipfile.ZipFile(nome_zip, "w", zipfile.ZIP_DEFLATED) as zipf:
    zipf.write("./Invio Email")

# Crea un oggetto MIMEMultipart per l'email
email = MIMEMultipart()
email["From"] = mittente
email["To"] = destinatario
email["Subject"] = oggetto

# Aggiungi il testo dell'email
email.attach(MIMEText(testo, "plain"))

# Aggiungi il file zip come allegato
with open(nome_zip, "rb") as file:
    allegato = MIMEBase("application", "zip")
    allegato.set_payload(file.read())
encoders.encode_base64(allegato)
allegato.add_header("Content-Disposition", f"attachment; filename= {nome_zip}")
email.attach(allegato)

# Invia l'email tramite il server SMTP di Gmail
with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
    server.login(mittente, password)
    server.sendmail(mittente, destinatario, email.as_string())

print("Email inviata con successo.")

# Rimuovi il file zip dopo l'invio dell'email
os.remove(nome_zip)
